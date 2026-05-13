# reporter/excel.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
import os

from data.config import DAYS, CHANGEOVER_DAYS
from data.shift_calculator import (
    active_buffets_by_day, accom_target_for_day,
)

# ── Colours ───────────────────────────────────────────────────────────────────
C = {
    "bg":     "0E0F11", "hdr":  "1E2023", "alt":   "161719",
    "gold":   "C8A96E", "white":"E8E9EB", "dim":   "6B6F76",
    "green":  "4ADE80", "amber":"FBBF24", "red":   "F87171",
    "shift":  "1E3A2F", "shift_t":"4ADE80",
    "rest":   "1E2023", "rest_t": "6B6F76",
    "hol":    "2E2418", "hol_t":  "C8A96E",
    "under":  "3A1A1A", "ok":     "1A3A26", "over": "3A2E0E",
    "accom":  "162030", "accom_t":"60A5FA",
    "chgov":  "2A1A30", "chgov_t":"C084FC",
    # Cross-venue: bar shift shown on a non-bar sheet (or vice-versa)
    "xbar":   "2E1800", "xbar_t": "FBBF24",
    # Cross-venue: buffet shift shown on another buffet/bar sheet
    "xbuf":   "0D2828", "xbuf_t": "2DD4BF",
}

def _f(c="E8E9EB", bold=False, sz=10, it=False):
    return Font(color=c, bold=bold, size=sz, name="Arial", italic=it)

def _fill(c):
    return PatternFill("solid", fgColor=c)

def _b():
    s = Side(style="thin", color="2A2C30")
    return Border(left=s, right=s, top=s, bottom=s)

def _a(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _hdr_row(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill(C["hdr"]); cell.font = _f(C["gold"], bold=True, sz=9)
        cell.border = _b(); cell.alignment = _a()

def _week_dates(week_start_str):
    d = datetime.strptime(week_start_str, "%Y-%m-%d").date()
    return [d + timedelta(days=i) for i in range(7)]


def export_to_excel(results, week_config, buffet_schedule, bar_schedule):
    os.makedirs("outputs", exist_ok=True)
    wb    = openpyxl.Workbook()
    dates = _week_dates(week_config["week_start"])

    # Summary
    _summary_sheet(wb.active, results, week_config)

    # Coverage report
    _coverage_sheet(wb.create_sheet("Coverage"), results, week_config)

    # Signed-form risk
    _risk_sheet(wb.create_sheet("Signed-Form Risk"), results)

    # Venue roster grids
    active = active_buffets_by_day(week_config)
    for vid in buffet_schedule:
        ws = wb.create_sheet(vid.replace("_"," ").title()[:31])
        _venue_roster(ws, vid, "buffet", results, week_config, dates, active)

    for vid in bar_schedule:
        ws = wb.create_sheet(vid.replace("_"," ").title()[:31])
        _venue_roster(ws, vid, "bar", results, week_config, dates, active)

    ws = wb.create_sheet("Accommodation")
    _venue_roster(ws, "accommodation", "accommodation",
                  results, week_config, dates, active)

    path = "outputs/roster_report.xlsx"
    wb.save(path)
    print(f"  Saved: {path}")


# ── Summary sheet ─────────────────────────────────────────────────────────────

def _summary_sheet(ws, results, wc):
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 24

    ws["A1"] = f"RESORT ROSTER — WEEK {wc['week_number']}"
    ws["A1"].font = _f(C["gold"], bold=True, sz=14)
    ws["A1"].fill = _fill(C["bg"]); ws["A1"].alignment = _a(h="left")
    ws.row_dimensions[1].height = 28

    ws["A2"] = (f"Generated {datetime.now().strftime('%a %d %b %Y · %H:%M')}   |   "
                f"Solver: {results.get('status', results['summary'].get('status','?'))}   |   "
                f"{results['solve_time']:.2f}s")
    ws["A2"].font = _f(C["dim"], sz=9)
    ws["A2"].fill = _fill(C["bg"]); ws["A2"].alignment = _a(h="left")

    s = results["summary"]
    rows = [
        ("Total staff",              str(s.get("total_staff",0)),            C["white"]),
        ("Total contracted hours",   f"{s.get('total_contracted_hrs',0)}h",  C["white"]),
        ("Total shift hours",        f"{s.get('total_shift_hrs',0)}h",       C["white"]),
        ("Total holiday hours",      f"{s.get('total_holiday_hrs',0)}h",     C["hol_t"]),
        ("Total allocated hours",    f"{s.get('total_allocated_hrs',0)}h",   C["green"]),
        ("Total hours gap",          f"{s.get('total_gap_hrs',0)}h",
         C["amber"] if s.get("total_gap_hrs",0) > 0 else C["green"]),
        ("Below contracted (staff)", str(s.get("below_contracted_count",0)),
         C["red"] if s.get("below_contracted_count",0) > 0 else C["green"]),
        ("Over contracted (staff)",  str(s.get("over_hours_count",0)),
         C["amber"] if s.get("over_hours_count",0) > 0 else C["green"]),
        ("Holiday days used",        str(s.get("holiday_days_total",0)),     C["white"]),
    ]
    for i, (lbl, val, col) in enumerate(rows, start=4):
        lc = ws.cell(row=i, column=1, value=lbl)
        lc.font=_f(); lc.fill=_fill(C["bg"]); lc.border=_b(); lc.alignment=_a(h="left")
        vc = ws.cell(row=i, column=2, value=val)
        vc.font=_f(col,bold=True); vc.fill=_fill(C["bg"]); vc.border=_b(); vc.alignment=_a()


# ── Coverage sheet ────────────────────────────────────────────────────────────

def _coverage_sheet(ws, results, week_config):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10
    for i in range(4, 11):
        ws.column_dimensions[get_column_letter(i)].width = 10

    headers = ["Venue", "Role", "Service"] + DAYS
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _hdr_row(ws, 1, len(headers))

    row = 2
    cov = results.get("coverage", {})
    for vid, days_data in cov.items():
        # Collect all (service, role) pairs across the week
        svc_roles = set()
        for ddata in days_data.values():
            for svc, roles in ddata.get("services", {}).items():
                for role in roles:
                    svc_roles.add((svc, role))

        for svc, role in sorted(svc_roles):
            lbl_v = vid.replace("_", " ").title()

            cell = ws.cell(row=row, column=1, value=lbl_v)
            cell.fill=_fill(C["bg"]); cell.font=_f(C["gold"],bold=True,sz=9)
            cell.border=_b(); cell.alignment=_a(h="left")

            cell = ws.cell(row=row, column=2,
                           value=role.replace("_"," ").title())
            cell.fill=_fill(C["bg"]); cell.font=_f(C["white"],sz=9)
            cell.border=_b(); cell.alignment=_a()

            cell = ws.cell(row=row, column=3,
                           value=svc.replace("_"," ").title())
            cell.fill=_fill(C["bg"]); cell.font=_f(C["dim"],sz=9)
            cell.border=_b(); cell.alignment=_a()

            for day in range(7):
                col   = day + 4
                ddata = days_data.get(day, {})
                rdata = ddata.get("services", {}).get(svc, {}).get(role, {})

                if not ddata.get("open"):
                    val = "—"; bg = C["rest"]; fg = C["dim"]
                elif not rdata:
                    val = "—"; bg = C["rest"]; fg = C["dim"]
                else:
                    actual = rdata.get("actual", 0)
                    target = rdata.get("target", 0)
                    gap    = rdata.get("gap", 0)
                    val    = f"{actual}/{target}"
                    if gap == 0:
                        bg = C["ok"];    fg = C["green"]
                    elif gap > 0:
                        bg = C["under"]; fg = C["red"]
                    else:
                        bg = C["over"];  fg = C["amber"]

                cell = ws.cell(row=row, column=col, value=val)
                cell.fill=_fill(bg); cell.font=_f(fg,sz=9,bold=True)
                cell.border=_b(); cell.alignment=_a()

            ws.row_dimensions[row].height = 18
            row += 1


# ── Risk sheet ────────────────────────────────────────────────────────────────

def _risk_sheet(ws, results):
    ws.sheet_view.showGridLines = False
    cols   = ["Name","Type","Department","Contracted h","Allocated h","Shortfall h","Action"]
    widths = [22, 10, 16, 14, 14, 14, 34]
    for i, (h, w) in enumerate(zip(cols, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        ws.cell(row=1, column=i, value=h)
    _hdr_row(ws, 1, len(cols))

    risk = results["summary"].get("below_contracted_staff", [])
    all_a = results.get("assignments", [])
    amap  = {a["staff_id"]: a for a in all_a}

    if not risk:
        c = ws.cell(row=2, column=1, value="No staff below contracted hours ✓")
        c.font=_f(C["green"],bold=True); c.fill=_fill(C["bg"]); return

    for r, p in enumerate(risk, start=2):
        a = next((x for x in all_a if x["name"]==p["name"]), {})
        for c, val in enumerate([
            p["name"],
            a.get("contract_type",""),
            a.get("department",""),
            p["contracted"], p["allocated"], p["shortfall"],
            "Obtain signed voluntary reduction form",
        ], 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill=_fill(C["under"]); cell.font=_f(C["red"])
            cell.border=_b(); cell.alignment=_a()
        ws.row_dimensions[r].height = 20


# ── Venue roster grid ─────────────────────────────────────────────────────────

def _venue_roster(ws, vid, venue_type, results, week_config, dates, active):
    ws.sheet_view.showGridLines = False
    name = vid.replace("_"," ").title()

    # Filter staff assigned to this venue
    all_a    = results.get("assignments", [])
    venue_staff = [
        a for a in all_a
        if any(
            sh.get("venue_id") == vid
            for ds in a["days"].values()
            for sh in ds
            if isinstance(sh, dict) and "venue_id" in sh
        )
    ]

    # Column layout: Name | Role | Day0..Day6 | Shift h | Hol h | Total h | Contract h | Var
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    for i in range(7):
        ws.column_dimensions[get_column_letter(i+3)].width = 18
    ws.column_dimensions[get_column_letter(10)].width = 10  # shift h
    ws.column_dimensions[get_column_letter(11)].width = 8   # hol h
    ws.column_dimensions[get_column_letter(12)].width = 10  # total h
    ws.column_dimensions[get_column_letter(13)].width = 11  # contract h
    ws.column_dimensions[get_column_letter(14)].width = 10  # var

    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = f"ROSTER — {name.upper()}"
    ws["A1"].font = _f(C["gold"], bold=True, sz=14)
    ws["A1"].fill = _fill(C["bg"]); ws["A1"].alignment = _a(h="left")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:L2")
    ws["A2"] = (f"Week {week_config['week_number']}  "
                f"{dates[0].strftime('%d %b')} – {dates[6].strftime('%d %b %Y')}  |  "
                f"{len(venue_staff)} staff")
    ws["A2"].font = _f(C["dim"], sz=9)
    ws["A2"].fill = _fill(C["bg"]); ws["A2"].alignment = _a(h="left")
    ws.row_dimensions[2].height = 16

    # Header row
    day_headers = [
        f"{DAYS[d]}\n{dates[d].strftime('%d %b')}"
        + (" ★" if d in CHANGEOVER_DAYS else "")
        for d in range(7)
    ]
    headers = ["Team Member","Role/Task"] + day_headers + \
              ["Shift\nh","Hol\nh","Total\nh","Contract\nh","Var\nh"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _hdr_row(ws, 3, len(headers))
    ws.row_dimensions[3].height = 30

    if not venue_staff:
        ws.merge_cells("A4:N4")
        ws["A4"] = "No staff allocated this week."
        ws["A4"].font = _f(C["dim"],it=True); ws["A4"].fill=_fill(C["bg"])
        return

    for r_idx, a in enumerate(venue_staff):
        row = r_idx + 4
        alt = C["alt"] if r_idx % 2 else C["bg"]
        ws.row_dimensions[row].height = 42
        is_tm = a["contract_type"] == "tm_plus"

        # Name
        nc = ws.cell(row=row, column=1, value=a["name"])
        nc.font  = _f(C["gold"] if is_tm else C["white"], bold=True)
        nc.fill  = _fill(alt); nc.border = _b(); nc.alignment = _a(h="left")

        # Role column — primary skill
        skill = next((sk for sk in a.get("skills",[])
                      if sk not in ("accommodation","accommodation_general",
                                    "bartender")), "")
        rc = ws.cell(row=row, column=2,
                     value=skill.replace("_"," ").title())
        rc.font = _f(C["dim"],sz=9); rc.fill=_fill(alt)
        rc.border=_b(); rc.alignment=_a()

        # Day cells
        for d_idx in range(7):
            col        = d_idx + 3
            day_shifts = a["days"].get(d_idx, [])
            is_holiday = any(sh.get("type") == "holiday" for sh in day_shifts)
            is_changeover = d_idx in CHANGEOVER_DAYS

            venue_shifts = [
                sh for sh in day_shifts
                if isinstance(sh, dict) and sh.get("venue_id") == vid
            ]
            other_shifts = [
                sh for sh in day_shifts
                if isinstance(sh, dict) and sh.get("venue_id")
                and sh.get("venue_id") != vid
            ]

            def _slabel(sh):
                task = sh.get("task", sh.get("role","").replace("_"," ").title())
                lbl  = sh.get("label","")
                return f"{task}\n{lbl}" if lbl else task

            def _vtype(sh):
                if sh.get("venue_id") == "accommodation":
                    return "accommodation"
                if sh.get("session"):   # bar shifts carry a session field
                    return "bar"
                return "buffet"

            def _cross_color(sh, day_idx):
                vt = _vtype(sh)
                if vt == "accommodation":
                    return (C["chgov"], C["chgov_t"]) if day_idx in CHANGEOVER_DAYS \
                           else (C["accom"], C["accom_t"])
                if vt == "bar":
                    return C["xbar"], C["xbar_t"]
                return C["xbuf"], C["xbuf_t"]

            if is_holiday:
                label = "HOLIDAY"
                bg = C["hol"]; fg = C["hol_t"]

            elif not venue_shifts and not other_shifts:
                label = "REST"
                bg = C["rest"]; fg = C["rest_t"]

            elif not venue_shifts:
                # Staff is working elsewhere this day — show where
                parts = []
                for sh in other_shifts:
                    vname = sh.get("venue_id","").replace("_"," ").title()
                    parts.append(f"[{vname}]\n{_slabel(sh)}")
                label = "\n".join(parts)
                # Color by the first other shift's venue type
                bg, fg = _cross_color(other_shifts[0], d_idx)

            else:
                # Working this venue; also flag any additional venues
                parts = [_slabel(sh) for sh in venue_shifts]
                for sh in other_shifts:
                    vname = sh.get("venue_id","").replace("_"," ").title()
                    parts.append(f"+{vname}\n{_slabel(sh)}")
                label = "\n".join(parts)
                if vid == "accommodation":
                    bg = C["chgov"] if is_changeover else C["accom"]
                    fg = C["chgov_t"] if is_changeover else C["accom_t"]
                else:
                    bg = C["shift"]; fg = C["shift_t"]

            cell = ws.cell(row=row, column=col, value=label)
            cell.fill=_fill(bg); cell.font=_f(fg,sz=8,bold=(bg==C["shift"]))
            cell.border=_b(); cell.alignment=_a(wrap=True)

        # Shift hours
        sc = ws.cell(row=row, column=10, value=a.get("shift_hours", a["allocated_hours"]))
        sc.font=_f(C["white"]); sc.fill=_fill(alt); sc.border=_b(); sc.alignment=_a()
        # Holiday hours
        hc2 = ws.cell(row=row, column=11, value=a.get("holiday_hours",0))
        hc2.font=_f(C["hol_t"]); hc2.fill=_fill(alt); hc2.border=_b(); hc2.alignment=_a()
        # Total (formula)
        tc2 = ws.cell(row=row, column=12, value=f"=J{row}+K{row}")
        tc2.font=_f(C["white"],bold=True); tc2.fill=_fill(alt)
        tc2.border=_b(); tc2.alignment=_a()
        # Contracted
        cc = ws.cell(row=row, column=13, value=a["contracted_hours"])
        cc.font=_f(C["dim"]); cc.fill=_fill(alt); cc.border=_b(); cc.alignment=_a()
        # Variance
        vc = ws.cell(row=row, column=14, value=f"=L{row}-M{row}")
        vc.fill=_fill(C["under"] if a["below_contracted"] else
                      C["over"]  if a["hours_over"]>0.1 else C["ok"])
        vc.font=_f(C["red"]   if a["below_contracted"] else
                   C["amber"] if a["hours_over"]>0.1 else C["green"], bold=True)
        vc.border=_b(); vc.alignment=_a()

    # Totals row
    tr = len(venue_staff) + 4
    ws.row_dimensions[tr].height = 22
    tc = ws.cell(row=tr, column=1, value="TOTAL")
    tc.font=_f(C["gold"],bold=True); tc.fill=_fill(C["hdr"])
    tc.border=_b(); tc.alignment=_a(h="left")
    for col in range(2, 10):
        ws.cell(row=tr,column=col,value="").fill=_fill(C["hdr"])
        ws.cell(row=tr,column=col).border=_b()
    # Shift h total
    ta = ws.cell(row=tr,column=10,value=f"=SUM(J4:J{tr-1})")
    ta.font=_f(C["white"],bold=True); ta.fill=_fill(C["hdr"])
    ta.border=_b(); ta.alignment=_a()
    # Holiday h total
    th = ws.cell(row=tr,column=11,value=f"=SUM(K4:K{tr-1})")
    th.font=_f(C["hol_t"]); th.fill=_fill(C["hdr"])
    th.border=_b(); th.alignment=_a()
    # Total h
    tt = ws.cell(row=tr,column=12,value=f"=SUM(L4:L{tr-1})")
    tt.font=_f(C["white"],bold=True); tt.fill=_fill(C["hdr"])
    tt.border=_b(); tt.alignment=_a()
    # Contract h total
    tb = ws.cell(row=tr,column=13,value=f"=SUM(M4:M{tr-1})")
    tb.font=_f(C["gold"]); tb.fill=_fill(C["hdr"])
    tb.border=_b(); tb.alignment=_a()
    # Variance total
    tv = ws.cell(row=tr,column=14,value=f"=L{tr}-M{tr}")
    tv.font=_f(C["green"],bold=True); tv.fill=_fill(C["ok"])
    tv.border=_b(); tv.alignment=_a()

    # Headcount per day — split by service
    if venue_type == "buffet":
        services = ["breakfast", "dinner"]
    else:
        services = ["total"]   # bars and accommodation show total headcount

    for svc_idx, svc in enumerate(services):
        hc  = tr + 1 + svc_idx
        ws.row_dimensions[hc].height = 16
        lbl = f"HC {svc.title()}" if len(services)>1 else "Headcount"
        ws.cell(row=hc,column=1,value=lbl).font=_f(C["dim"],sz=8,it=True)
        ws.cell(row=hc,column=1).fill=_fill(C["bg"])
        ws.cell(row=hc,column=1).border=_b()
        ws.cell(row=hc,column=2,value="").fill=_fill(C["bg"])
        ws.cell(row=hc,column=2).border=_b()

        for d_idx in range(7):
            if svc == "total":
                count = sum(
                    1 for a in venue_staff
                    if any(sh.get("venue_id")==vid
                           for sh in a["days"].get(d_idx,[])
                           if isinstance(sh,dict))
                )
                # For bars/accommodation: look up any service key
                cov_day = results.get("coverage",{}).get(vid,{}).get(d_idx,{})
                tgt = 0
                for svc_data in cov_day.get("services",{}).values():
                    tgt += sum(r.get("target",0) for r in svc_data.values()
                               if isinstance(r,dict))
            else:
                count = sum(
                    1 for a in venue_staff
                    if any(sh.get("venue_id")==vid
                           and sh.get("service")==svc
                           for sh in a["days"].get(d_idx,[])
                           if isinstance(sh,dict))
                )
                # Get target from coverage
                cov_day = results.get("coverage",{}).get(vid,{}).get(d_idx,{})
                cov_svc = cov_day.get("services",{}).get(svc,{})
                tgt     = sum(r.get("target",0) for r in cov_svc.values()
                              if isinstance(r,dict)) if cov_svc else 0

            label = f"{count}" if tgt==0 else f"{count}/{tgt}"
            ok    = count >= tgt if tgt > 0 else count > 0

            cell = ws.cell(row=hc,column=d_idx+3,value=label)
            cell.fill  = _fill(C["bg"])
            cell.font  = _f(C["green"] if ok else C["amber"],sz=8,bold=True)
            cell.border= _b(); cell.alignment=_a()

        for col in range(10,13):
            ws.cell(row=hc,column=col,value="").fill=_fill(C["bg"])
            ws.cell(row=hc,column=col).border=_b()

    last_hc = tr + len(services)

    # Legend
    lr = last_hc + 2
    ws.cell(row=lr,column=1,value="LEGEND").font=_f(C["dim"],sz=8,bold=True)
    ws.cell(row=lr,column=1).fill=_fill(C["bg"])
    for i,(lbl,bg,fg) in enumerate([
        ("Shift (task + time)", C["shift"],  C["shift_t"]),
        ("Accommodation daily", C["accom"],  C["accom_t"]),
        ("Changeover day",      C["chgov"],  C["chgov_t"]),
        ("REST",                C["rest"],   C["rest_t"]),
        ("HOLIDAY",             C["hol"],    C["hol_t"]),
        ("Gold name = TM+",     C["bg"],     C["gold"]),
    ], start=2):
        cell = ws.cell(row=lr, column=i, value=lbl)
        cell.fill=_fill(bg); cell.font=_f(fg,sz=8)
        cell.border=_b(); cell.alignment=_a()
