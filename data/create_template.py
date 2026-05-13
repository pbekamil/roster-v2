# =============================================================================
# data/create_template.py
# Generates inputs/week_template.xlsx pre-populated with current roster data.
# Run once (or whenever the roster structure changes):
#   python data/create_template.py
# Then copy the template each week:
#   cp inputs/week_template.xlsx inputs/week22.xlsx
# =============================================================================

import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from data.sample_data import STAFF, WEEK_CONFIG, BUFFET_SCHEDULE
from data.bars_data import BAR_SCHEDULE, DAILY_HOURS

_DAY_NAMES = ["Fri", "Sat", "Sun", "Mon", "Tue", "Wed", "Thu"]

# ── Style helpers ─────────────────────────────────────────────────────────────

_HDR_FILL  = PatternFill("solid", fgColor="2F5597")
_HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
_KEY_FILL  = PatternFill("solid", fgColor="D9E1F2")
_KEY_FONT  = Font(bold=True, size=10)
_THIN      = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _hdr(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font  = _HDR_FONT
    c.fill  = _HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _THIN
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def _key(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font  = _KEY_FONT
    c.fill  = _KEY_FILL
    c.border = _THIN
    return c


def _val(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.border = _THIN
    c.alignment = Alignment(horizontal="left")
    return c


def _dv(formula1, sqref, allow_blank=True, prompt=""):
    dv = DataValidation(type="list", formula1=formula1,
                        allow_blank=allow_blank, showInputMessage=True,
                        showErrorMessage=True, prompt=prompt)
    dv.sqref = sqref
    return dv


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_week_config(wb, wc):
    ws = wb.create_sheet("WeekConfig")
    ws.sheet_view.showGridLines = False

    rows = [
        ("week_number",            wc["week_number"]),
        ("week_start",             wc["week_start"]),
        ("fri_premium_guests",     wc["fri_batch"]["premium_guests"]),
        ("fri_food_court_guests",  wc["fri_batch"]["food_court_guests"]),
        ("fri_total_guests",       wc["fri_batch"]["total_guests"]),
        ("mon_premium_guests",     wc["mon_batch"]["premium_guests"]),
        ("mon_food_court_guests",  wc["mon_batch"]["food_court_guests"]),
        ("mon_total_guests",       wc["mon_batch"]["total_guests"]),
        ("fri_arrivals_override",  wc.get("accom_override", {}).get("fri_arrivals") or ""),
        ("mon_arrivals_override",  wc.get("accom_override", {}).get("mon_arrivals") or ""),
    ]

    _hdr(ws, 1, 1, "Key", width=28)
    _hdr(ws, 1, 2, "Value", width=20)

    for i, (k, v) in enumerate(rows, start=2):
        _key(ws, i, 1, k)
        _val(ws, i, 2, v)

    ws.row_dimensions[1].height = 18


def _build_staff(wb, staff):
    ws = wb.create_sheet("Staff")
    ws.sheet_view.showGridLines = False

    headers = [
        ("name", 24), ("contract_type", 14), ("tm_plus_type", 14),
        ("department", 14), ("home_venues", 20), ("eligible_venues", 36),
        ("contracted_hours", 14), ("hourly_rate", 12),
        ("holiday_remaining", 14), ("skills", 30),
    ]
    for col, (h, w) in enumerate(headers, start=1):
        _hdr(ws, 1, col, h, width=w)

    # Dropdowns
    n_rows = len(staff) + 50  # extra rows for new starters
    ws.add_data_validation(_dv('"normal,tm_plus"',         f"B2:B{n_rows}", prompt="normal or tm_plus"))
    ws.add_data_validation(_dv('"tm_plus_1,tm_plus_2"',    f"C2:C{n_rows}", prompt="blank / tm_plus_1 / tm_plus_2"))
    ws.add_data_validation(_dv('"buffets,bars,accommodation"', f"D2:D{n_rows}"))

    seen = set()
    for r, s in enumerate(staff, start=2):
        if s["name"].startswith("Accom "):
            # Collapse accommodation into one placeholder row
            if "accom_placeholder" not in seen:
                seen.add("accom_placeholder")
                accom_count = sum(1 for x in staff if x["name"].startswith("Accom "))
                ws.cell(row=r, column=1, value="Accom Team").border = _THIN
                ws.cell(row=r, column=2, value="normal").border      = _THIN
                ws.cell(row=r, column=3, value="").border             = _THIN
                ws.cell(row=r, column=4, value="accommodation").border = _THIN
                ws.cell(row=r, column=5, value="accommodation").border = _THIN
                ws.cell(row=r, column=6, value="accommodation").border = _THIN
                ws.cell(row=r, column=7, value=35.0).border           = _THIN
                ws.cell(row=r, column=8, value=11.00).border          = _THIN
                ws.cell(row=r, column=9, value=10).border             = _THIN
                ws.cell(row=r, column=10, value="").border            = _THIN
                # Add count column
                ws.cell(row=r, column=11, value=accom_count)
                if ws.cell(row=1, column=11).value is None:
                    _hdr(ws, 1, 11, "count (accom only)", width=18)
            continue

        skills_raw = [sk for sk in s["skills"] if sk not in ("floor", "setup", "accommodation")]
        row_vals = [
            s["name"],
            s["contract_type"],
            s.get("tm_plus_type") or "",
            s["department"],
            ",".join(s["home_venues"]),
            ",".join(v for v in s["eligible_venues"] if v != "accommodation"),
            s["contracted_hours"],
            s["hourly_rate"],
            s.get("holiday_remaining", 10),
            ",".join(skills_raw),
        ]
        for col, val in enumerate(row_vals, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = _THIN
            c.alignment = Alignment(horizontal="left")


def _build_days_off(wb, staff):
    ws = wb.create_sheet("DaysOff")
    ws.sheet_view.showGridLines = False

    _hdr(ws, 1, 1, "name", width=28)
    _hdr(ws, 1, 2, "day",  width=10)

    ws.add_data_validation(_dv('"Fri,Sat,Sun,Mon,Tue,Wed,Thu"', "B2:B500",
                               prompt="Day of week"))

    names = sorted({s["name"] for s in staff
                    if not s["name"].startswith("Accom ")})
    dv_names = DataValidation(
        type="list",
        formula1='"' + ",".join(names[:100]) + '"',  # first 100 only (Excel limit)
        allow_blank=True, showInputMessage=True,
    )
    ws.add_data_validation(dv_names)
    dv_names.sqref = "A2:A500"

    # Pre-populate with any existing days off from sample data
    row = 2
    for s in staff:
        if s["name"].startswith("Accom "):
            continue
        for d in s.get("approved_days_off", []):
            ws.cell(row=row, column=1, value=s["name"]).border = _THIN
            ws.cell(row=row, column=2, value=_DAY_NAMES[d]).border = _THIN
            row += 1


def _build_daily_hours(wb, daily_hours):
    ws = wb.create_sheet("DailyHours")
    ws.sheet_view.showGridLines = False

    _hdr(ws, 1, 1, "venue", width=18)
    for col, day in enumerate(_DAY_NAMES, start=2):
        _hdr(ws, 1, col, day, width=8)

    for r, (venue, hours) in enumerate(daily_hours.items(), start=2):
        ws.cell(row=r, column=1, value=venue).border = _THIN
        for col, h in enumerate(hours, start=2):
            c = ws.cell(row=r, column=col, value=h)
            c.border = _THIN
            c.alignment = Alignment(horizontal="center")


def _build_bar_schedule(wb, bar_schedule):
    ws = wb.create_sheet("BarSchedule")
    ws.sheet_view.showGridLines = False

    for col, (h, w) in enumerate(
        [("venue",14),("day",8),("session",14),("open",10),("close",10)], start=1
    ):
        _hdr(ws, 1, col, h, width=w)

    ws.add_data_validation(_dv('"Fri,Sat,Sun,Mon,Tue,Wed,Thu"', "B2:B500"))

    row = 2
    for venue, sessions in bar_schedule.items():
        for svc in sorted(sessions, key=lambda s: (s["day"], s["session"])):
            ws.cell(row=row, column=1, value=venue).border      = _THIN
            ws.cell(row=row, column=2, value=_DAY_NAMES[svc["day"]]).border = _THIN
            ws.cell(row=row, column=3, value=svc["session"]).border          = _THIN
            ws.cell(row=row, column=4, value=svc["open"]).border             = _THIN
            ws.cell(row=row, column=5, value=svc["close"]).border            = _THIN
            row += 1


def _build_buffet_schedule(wb, buffet_schedule):
    ws = wb.create_sheet("BuffetSchedule")
    ws.sheet_view.showGridLines = False

    for col, (h, w) in enumerate(
        [("venue",16),("day",8),("service",12),("open",10),("close",10)], start=1
    ):
        _hdr(ws, 1, col, h, width=w)

    ws.add_data_validation(_dv('"Fri,Sat,Sun,Mon,Tue,Wed,Thu"', "B2:B300"))
    ws.add_data_validation(_dv('"breakfast,dinner"', "C2:C300"))

    row = 2
    for venue, svcs in buffet_schedule.items():
        for svc in sorted(svcs, key=lambda s: (s["day"], s["service"])):
            ws.cell(row=row, column=1, value=venue).border      = _THIN
            ws.cell(row=row, column=2, value=_DAY_NAMES[svc["day"]]).border = _THIN
            ws.cell(row=row, column=3, value=svc["service"]).border          = _THIN
            ws.cell(row=row, column=4, value=svc["open"]).border             = _THIN
            ws.cell(row=row, column=5, value=svc["close"]).border            = _THIN
            row += 1


# ── Main ──────────────────────────────────────────────────────────────────────

def create_template(out_path="inputs/week_template.xlsx"):
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    _build_week_config(wb, WEEK_CONFIG)
    _build_staff(wb, STAFF)
    _build_days_off(wb, STAFF)
    _build_daily_hours(wb, DAILY_HOURS)
    _build_bar_schedule(wb, BAR_SCHEDULE)
    _build_buffet_schedule(wb, BUFFET_SCHEDULE)

    wb.save(out_path)
    print(f"  Template written to {out_path}")
    print(f"  Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "inputs/week_template.xlsx"
    create_template(path)
